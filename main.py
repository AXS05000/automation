from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.label import Label
import time


class MyApp(App):
    data_from_excel = []
    

    def build(self):
        self.title = 'Automation'

        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        import_excel_button = Button(text="Importar Planilha", on_press=self.show_filechooser)
        open_browser_button = Button(text="Abrir Navegador", on_press=self.open_browser)
        start_loop_button = Button(text="Iniciar Loop", on_press=self.start_loop)

        layout.add_widget(import_excel_button)
        layout.add_widget(open_browser_button)
        layout.add_widget(start_loop_button)
        return layout

    def show_filechooser(self, instance):
        self.filechooser = FileChooserIconView(filters=['*.xlsx'])
        self.filechooser.bind(on_submit=self.import_excel)
        self.popup = Popup(title="Selecione a Planilha", content=self.filechooser, size_hint=(0.9, 0.9))
        self.popup.open()

    def import_excel(self, instance, selection, touch):
        if not selection:
            return
        file_path = selection[0]

        try:
            # Carregar os dados da planilha
            wb = load_workbook(filename=file_path)
            ws = wb.active

            # Ler os dados da planilha para uma lista de dicionários
            for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorar o cabeçalho
                data = {
                    'jornadaRegraIdValue': row[0],
                    'dataInicioValue': row[1],
                    'usuarioIdValue': row[2]
                }
                self.data_from_excel.append(data)
            
            # Fechar o popup após a importação
            self.popup.dismiss()
            
            # Mensagem de confirmação
            confirmation = Popup(title="Concluído", content=Label(text="Planilha importada com sucesso!"), size_hint=(0.6, 0.4))
            confirmation.open()

        except Exception as e:
            error_msg = f"Erro ao importar a planilha: {str(e)}"
            error_popup = Popup(title="Erro", content=Label(text=error_msg), size_hint=(0.8, 0.6))
            error_popup.open()


    def configure_firefox(self):
        firefox_options = webdriver.FirefoxOptions()
        firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return firefox_options

    def open_browser(self, instance):
        firefox_options = self.configure_firefox()
        self.browser = webdriver.Firefox(options=firefox_options)
        self.browser.get('https://app.folhacerta.com/admin/login')

    def navigate_and_execute(self, data):
        time.sleep(5)
        self.browser.get('https://app.folhacerta.com/admin/JornadaRegraAlocacao?jornadaRegraAlocacao.Id=0')
        element_present = EC.presence_of_element_located((By.ID, 'content'))
        WebDriverWait(self.browser, 60).until(element_present)

        time.sleep(5)
        
        script = f"""
        var jornadaRegraIdValue = "{data['jornadaRegraIdValue']}";
        var dataInicioValue = "{data['dataInicioValue']}";
        var usuarioIdValue = "{data['usuarioIdValue']}";

        document.querySelector('[name="jornadaRegra_id"]').value = jornadaRegraIdValue;
        document.querySelector('[name="jornadaRegraAlocacao.DataInicio"]').value = dataInicioValue;

        var usuarioSelect = document.querySelector('#select-multi-usuarios');
        for (var i = 0; i < usuarioSelect.options.length; i++) {{
            if (usuarioSelect.options[i].value === usuarioIdValue) {{
                usuarioSelect.options[i].selected = true;
                break;
            }}
        }}

        document.querySelector('form.form-horizontal').submit();
        """
        self.browser.execute_script(script)

    def start_loop(self, instance):
        for data in self.data_from_excel:
            self.navigate_and_execute(data)

if __name__ == '__main__':
    MyApp().run()
